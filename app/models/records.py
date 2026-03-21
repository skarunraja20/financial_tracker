"""
Records model: investment holder/nominee details, protection records
(insurance, emergency fund, will) and personal contacts.
"""

from datetime import datetime
from app.core.database import get_connection

# ── Reference look-up dictionaries ────────────────────────────────────────────

PROTECTION_TYPES = {
    "emergency_fund":    "Emergency Fund",
    "health_insurance":  "Health Insurance",
    "term_insurance":    "Term Insurance",
    "life_insurance":    "Life Insurance",
    "vehicle_insurance": "Vehicle Insurance",
    "critical_illness":  "Critical Illness Ins.",
    "will":              "Will / Testament",
    "other":             "Other",
}

PREMIUM_FREQUENCIES = {
    "monthly":     "Monthly",
    "quarterly":   "Quarterly",
    "half_yearly": "Half-Yearly",
    "annual":      "Annual",
    "one_time":    "One-Time",
    "na":          "N/A",
}

CONTACT_TYPES = {
    "emergency":         "Emergency Contact",
    "advocate":          "Advocate / Lawyer",
    "financial_advisor": "Financial Advisor",
    "ca_tax_advisor":    "CA / Tax Advisor",
    "doctor":            "Doctor",
    "banker":            "Banker",
    "insurance_agent":   "Insurance Agent",
    "other":             "Other",
}

# Asset class labels for each asset_type key
_ASSET_CLASS = {
    "pf":           "Debt - PF",
    "ppf":          "Debt - PPF",
    "nps":          "Debt - NPS",
    "fd":           "Debt - Fixed Deposit",
    "bonds":        "Debt - Bond",
    "debt_mf":      "Debt - Mutual Fund",
    "equity_mf":    "Equity - Mutual Fund",
    "stocks":       "Equity - Stock",
    "gold_mf":      "Gold - Mutual Fund",
    "sgb":          "Gold - SGB",
    "real_estate":  "Real Estate",
}


# ── Investment Records ─────────────────────────────────────────────────────────

def get_all_investment_records() -> list[dict]:
    """
    Returns every active investment asset merged with its investment_record row
    (if any) and goal tags.  Assets without a record row appear with blank
    holder/nominee fields — ready for the user to fill in.
    """
    conn = get_connection()
    try:
        # UNION ALL across every asset table to get a flat list
        assets_sql = """
            SELECT asset_type, asset_id, asset_name, acct_hint
            FROM (
                SELECT 'pf' AS asset_type, 1 AS asset_id,
                       'Provident Fund (PF)' AS asset_name,
                       account_number AS acct_hint
                FROM   pf_account
                LIMIT  1
            )

            UNION ALL

            SELECT asset_type, asset_id, asset_name, acct_hint FROM (
                SELECT 'ppf' AS asset_type, 1 AS asset_id,
                       'Public Provident Fund (PPF)' AS asset_name,
                       account_number AS acct_hint
                FROM   ppf_account
                LIMIT  1
            )

            UNION ALL

            SELECT asset_type, asset_id, asset_name, acct_hint FROM (
                SELECT 'nps' AS asset_type, 1 AS asset_id,
                       'National Pension System (NPS)' AS asset_name,
                       pran_number AS acct_hint
                FROM   nps_account
                LIMIT  1
            )

            UNION ALL

            SELECT 'fd', id, bank_name, fd_number
            FROM   fixed_deposits
            WHERE  is_active = 1

            UNION ALL

            SELECT 'bonds', id, bond_name, NULL
            FROM   bonds
            WHERE  is_active = 1

            UNION ALL

            SELECT 'debt_mf', id, fund_name, folio_number
            FROM   mutual_funds
            WHERE  fund_category = 'debt' AND is_active = 1

            UNION ALL

            SELECT 'equity_mf', id, fund_name, folio_number
            FROM   mutual_funds
            WHERE  fund_category = 'equity' AND is_active = 1

            UNION ALL

            SELECT 'gold_mf', id, fund_name, folio_number
            FROM   mutual_funds
            WHERE  fund_category = 'gold' AND is_active = 1

            UNION ALL

            SELECT 'stocks', id, company_name, demat_account
            FROM   stocks
            WHERE  is_active = 1

            UNION ALL

            SELECT 'sgb', id, series_name, NULL
            FROM   sgb
            WHERE  is_active = 1

            UNION ALL

            SELECT 'real_estate', id, property_name, NULL
            FROM   real_estate
        """
        assets = [dict(r) for r in conn.execute(assets_sql).fetchall()]

        # Index existing investment_records by (asset_type, asset_id)
        inv_recs: dict[tuple, dict] = {}
        for row in conn.execute("SELECT * FROM investment_records").fetchall():
            inv_recs[(row["asset_type"], row["asset_id"])] = dict(row)

        # Index goal names by (asset_type, asset_id)
        goal_tags: dict[tuple, list[str]] = {}
        for row in conn.execute("""
            SELECT ag.asset_type, ag.asset_id, g.name
            FROM   asset_goals ag
            JOIN   goals g ON g.id = ag.goal_id AND g.is_active = 1
        """).fetchall():
            key = (row["asset_type"], row["asset_id"])
            goal_tags.setdefault(key, []).append(row["name"])

        result = []
        for a in assets:
            key   = (a["asset_type"], a["asset_id"])
            rec   = inv_recs.get(key, {})
            goals = ", ".join(goal_tags.get(key, [])) or "—"

            # Prefer the saved account_folio_number; fall back to the hint from the asset table
            acct = rec.get("account_folio_number") or a.get("acct_hint") or ""

            result.append({
                "asset_type":           a["asset_type"],
                "asset_id":             a["asset_id"],
                "asset_name":           a["asset_name"],
                "asset_class":          _ASSET_CLASS.get(a["asset_type"], "—"),
                "acct_hint":            a.get("acct_hint") or "",
                "account_folio_number": acct,
                "first_holder":         rec.get("first_holder")  or "",
                "second_holder":        rec.get("second_holder") or "",
                "nominee_1_name":       rec.get("nominee_1_name") or "",
                "nominee_1_pct":        rec.get("nominee_1_pct")  or 0.0,
                "nominee_2_name":       rec.get("nominee_2_name") or "",
                "nominee_2_pct":        rec.get("nominee_2_pct")  or 0.0,
                "goals":                goals,
                "notes":                rec.get("notes") or "",
            })

        return result
    finally:
        conn.close()


def upsert_investment_record(asset_type: str, asset_id: int, **fields) -> None:
    """Insert or update the investment record for a given asset."""
    now = datetime.utcnow().isoformat()
    conn = get_connection()
    try:
        conn.execute("""
            INSERT INTO investment_records
                (asset_type, asset_id, account_folio_number, first_holder,
                 second_holder, nominee_1_name, nominee_1_pct,
                 nominee_2_name, nominee_2_pct, notes, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(asset_type, asset_id) DO UPDATE SET
                account_folio_number = excluded.account_folio_number,
                first_holder         = excluded.first_holder,
                second_holder        = excluded.second_holder,
                nominee_1_name       = excluded.nominee_1_name,
                nominee_1_pct        = excluded.nominee_1_pct,
                nominee_2_name       = excluded.nominee_2_name,
                nominee_2_pct        = excluded.nominee_2_pct,
                notes                = excluded.notes,
                updated_at           = excluded.updated_at
        """, (
            asset_type, asset_id,
            fields.get("account_folio_number", ""),
            fields.get("first_holder", ""),
            fields.get("second_holder", ""),
            fields.get("nominee_1_name", ""),
            fields.get("nominee_1_pct") or 0.0,
            fields.get("nominee_2_name", ""),
            fields.get("nominee_2_pct") or 0.0,
            fields.get("notes", ""),
            now,
        ))
        conn.commit()
    finally:
        conn.close()


# ── Protection Records ─────────────────────────────────────────────────────────

def get_all_protection_records() -> list[dict]:
    conn = get_connection()
    try:
        rows = conn.execute("""
            SELECT * FROM protection_records
            WHERE  is_active = 1
            ORDER  BY record_type, provider
        """).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def add_protection_record(**fields) -> int:
    now = datetime.utcnow().isoformat()
    conn = get_connection()
    try:
        cur = conn.execute("""
            INSERT INTO protection_records
                (record_type, provider, policy_number, coverage_amount,
                 premium_amount, premium_frequency, start_date, end_date,
                 nominee, notes, is_active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
        """, (
            fields.get("record_type", "other"),
            fields.get("provider", ""),
            fields.get("policy_number", ""),
            fields.get("coverage_amount") or None,
            fields.get("premium_amount")  or None,
            fields.get("premium_frequency", "annual"),
            fields.get("start_date", ""),
            fields.get("end_date", ""),
            fields.get("nominee", ""),
            fields.get("notes", ""),
            now, now,
        ))
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def update_protection_record(rec_id: int, **fields) -> None:
    now = datetime.utcnow().isoformat()
    conn = get_connection()
    try:
        conn.execute("""
            UPDATE protection_records
            SET record_type=?, provider=?, policy_number=?,
                coverage_amount=?, premium_amount=?, premium_frequency=?,
                start_date=?, end_date=?, nominee=?, notes=?, updated_at=?
            WHERE id=?
        """, (
            fields.get("record_type", "other"),
            fields.get("provider", ""),
            fields.get("policy_number", ""),
            fields.get("coverage_amount") or None,
            fields.get("premium_amount")  or None,
            fields.get("premium_frequency", "annual"),
            fields.get("start_date", ""),
            fields.get("end_date", ""),
            fields.get("nominee", ""),
            fields.get("notes", ""),
            now, rec_id,
        ))
        conn.commit()
    finally:
        conn.close()


def delete_protection_record(rec_id: int) -> None:
    """Soft-delete: marks is_active = 0 to preserve history."""
    conn = get_connection()
    try:
        conn.execute(
            "UPDATE protection_records SET is_active=0 WHERE id=?", (rec_id,)
        )
        conn.commit()
    finally:
        conn.close()


# ── Contact Records ────────────────────────────────────────────────────────────

def get_all_contact_records() -> list[dict]:
    conn = get_connection()
    try:
        rows = conn.execute("""
            SELECT * FROM contact_records
            ORDER  BY contact_type, name
        """).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def add_contact_record(**fields) -> int:
    now = datetime.utcnow().isoformat()
    conn = get_connection()
    try:
        cur = conn.execute("""
            INSERT INTO contact_records
                (contact_type, name, relationship, phone,
                 email, address, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            fields.get("contact_type", "other"),
            fields.get("name", ""),
            fields.get("relationship", ""),
            fields.get("phone", ""),
            fields.get("email", ""),
            fields.get("address", ""),
            fields.get("notes", ""),
            now, now,
        ))
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def update_contact_record(rec_id: int, **fields) -> None:
    now = datetime.utcnow().isoformat()
    conn = get_connection()
    try:
        conn.execute("""
            UPDATE contact_records
            SET contact_type=?, name=?, relationship=?, phone=?,
                email=?, address=?, notes=?, updated_at=?
            WHERE id=?
        """, (
            fields.get("contact_type", "other"),
            fields.get("name", ""),
            fields.get("relationship", ""),
            fields.get("phone", ""),
            fields.get("email", ""),
            fields.get("address", ""),
            fields.get("notes", ""),
            now, rec_id,
        ))
        conn.commit()
    finally:
        conn.close()


def delete_contact_record(rec_id: int) -> None:
    """Hard-delete a contact record."""
    conn = get_connection()
    try:
        conn.execute("DELETE FROM contact_records WHERE id=?", (rec_id,))
        conn.commit()
    finally:
        conn.close()


# ── Excel Export ───────────────────────────────────────────────────────────────

def export_records_to_excel(file_path: str) -> None:
    """
    Write a 3-sheet Excel workbook:
      Sheet 1 — Investments  (assets + holder / nominee details)
      Sheet 2 — Protection & Insurance
      Sheet 3 — Contacts
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── Shared style helpers ───────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")   # deep navy header
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)  # white bold header text
    DATA_FONT = Font(color="000000", size=9)              # black data text
    ROW_FILL  = PatternFill("solid", fgColor="FFFFFF")    # plain white rows
    ALT_FILL  = PatternFill("solid", fgColor="D6EAF8")    # pale blue alternate rows
    WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, headers: list[str], col_widths: list[int]):
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = w

    def write_row(ws, row_idx: int, values: list, alt: bool = False):
        fill = ALT_FILL if alt else ROW_FILL   # always apply a fill
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.alignment = WRAP_ALIGN
        ws.row_dimensions[row_idx].height = 16

    # ── Sheet 1: Investments ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Investments"
    ws1.sheet_view.showGridLines = False

    inv_headers = [
        "Asset Name", "Asset Class", "Account / Folio No.",
        "1st Holder", "2nd Holder",
        "Nominee 1 Name", "Nom 1 %",
        "Nominee 2 Name", "Nom 2 %",
        "Goals Tagged", "Notes",
    ]
    inv_widths = [28, 22, 22, 18, 18, 18, 8, 18, 8, 24, 30]
    style_header(ws1, inv_headers, inv_widths)

    inv_records = get_all_investment_records()
    for i, r in enumerate(inv_records):
        n1pct = f"{int(r['nominee_1_pct'])}%" if r["nominee_1_pct"] else ""
        n2pct = f"{int(r['nominee_2_pct'])}%" if r["nominee_2_pct"] else ""
        write_row(ws1, i + 2, [
            r["asset_name"], r["asset_class"], r["account_folio_number"],
            r["first_holder"], r["second_holder"],
            r["nominee_1_name"], n1pct,
            r["nominee_2_name"], n2pct,
            r["goals"], r["notes"],
        ], alt=(i % 2 == 1))

    # ── Sheet 2: Protection & Insurance ───────────────────────────────────────
    ws2 = wb.create_sheet("Protection & Insurance")
    ws2.sheet_view.showGridLines = False

    prot_headers = [
        "Type", "Provider / Institution", "Policy / Account No.",
        "Coverage Amt (₹)", "Premium Amt (₹)", "Frequency",
        "Start Date", "End / Renewal Date", "Nominee", "Notes",
    ]
    prot_widths = [22, 24, 22, 18, 16, 13, 12, 18, 18, 30]
    style_header(ws2, prot_headers, prot_widths)

    prot_records = get_all_protection_records()
    for i, r in enumerate(prot_records):
        write_row(ws2, i + 2, [
            PROTECTION_TYPES.get(r["record_type"], r["record_type"]),
            r["provider"] or "",
            r["policy_number"] or "",
            r["coverage_amount"] or "",
            r["premium_amount"] or "",
            PREMIUM_FREQUENCIES.get(r["premium_frequency"], r["premium_frequency"] or ""),
            r["start_date"] or "",
            r["end_date"] or "",
            r["nominee"] or "",
            r["notes"] or "",
        ], alt=(i % 2 == 1))

    # ── Sheet 3: Contacts ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Contacts")
    ws3.sheet_view.showGridLines = False

    cont_headers = [
        "Type", "Name", "Relationship",
        "Phone", "Email", "Address / Notes",
    ]
    cont_widths = [20, 22, 18, 16, 26, 40]
    style_header(ws3, cont_headers, cont_widths)

    cont_records = get_all_contact_records()
    for i, r in enumerate(cont_records):
        addr_notes = "\n".join(filter(None, [r.get("address") or "", r.get("notes") or ""]))
        write_row(ws3, i + 2, [
            CONTACT_TYPES.get(r["contact_type"], r["contact_type"]),
            r["name"],
            r["relationship"] or "",
            r["phone"] or "",
            r["email"] or "",
            addr_notes,
        ], alt=(i % 2 == 1))

    wb.save(file_path)
